import requests
import openpyxl
from datetime import datetime
from xml.etree import ElementTree as ET
import os


def get_rates():
    url = "https://www.cbr.ru/scripts/XML_daily.asp"
    response = requests.get(url, headers={'User-Agent': 'Mozilla/5.0'})
    response.encoding = 'windows-1251'
    
    root = ET.fromstring(response.text)
    
    rates = {}
    wanted = {'USD', 'EUR', 'CNY'}
    
    for valute in root.findall('Valute'):
        char_code = valute.find('CharCode').text
        if char_code in wanted:
            name = valute.find('Name').text
            value = valute.find('Value').text.replace(',', '.')
            nominal = valute.find('Nominal').text
            rates[char_code] = {
                'name': name,
                'value': round(float(value) / int(nominal), 4),
                'nominal': int(nominal),
            }
    
    return rates
def save_rates(rates):
    filename = os.path.expanduser('~/currency-tracker/history.xlsx')
    
    # Если файл уже есть — открываем, если нет — создаём
    if os.path.exists(filename):
        wb = openpyxl.load_workbook(filename)
        ws = wb.active
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "История курсов"
        # Заголовки только при создании
        ws.append(['Дата', 'USD', 'EUR', 'CNY'])
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 12
        ws.column_dimensions['C'].width = 12
        ws.column_dimensions['D'].width = 12

    # Добавляем новую строку
    now = datetime.now().strftime('%d.%m.%Y %H:%M')
    usd = rates.get('USD', {}).get('value', '-')
    eur = rates.get('EUR', {}).get('value', '-')
    cny = rates.get('CNY', {}).get('value', '-')
    ws.append([now, usd, eur, cny])

    wb.save(filename)
    return filename


def show_rates(rates):
    print(f"\nКурсы ЦБ РФ на {datetime.now().strftime('%d.%m.%Y')}:")
    print("-" * 35)
    for code, data in rates.items():
        print(f"{code} ({data['name']}): {data['value']} ₽")
    print("-" * 35)

def check_change(rates):
    filename = os.path.expanduser('~/currency-tracker/history.xlsx')
    
    if not os.path.exists(filename):
        return
    
    wb = openpyxl.load_workbook(filename)
    ws = wb.active
    
    # Нужно минимум 2 строки данных (+ заголовок = 3 строки)
    if ws.max_row < 3:
        return
    
    # Берём предпоследнюю строку
    prev_row = ws[ws.max_row - 1]
    prev_usd = prev_row[1].value
    curr_usd = rates.get('USD', {}).get('value')
    
    if prev_usd and curr_usd:
        change = ((curr_usd - prev_usd) / prev_usd) * 100
        if abs(change) >= 0.1:
            direction = "вырос" if change > 0 else "упал"
            print(f"\n⚠️  USD {direction} на {abs(change):.2f}% с последней проверки!")

if __name__ == '__main__':
    print("Получаем курсы с сайта ЦБ РФ...")
    rates = get_rates()
    show_rates(rates)
    check_change(rates)
    filename = save_rates(rates)
    print(f"\nИстория сохранена в: {filename}")